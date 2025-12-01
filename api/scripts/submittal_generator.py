"""
Submittal Log Generator - Analyzes specs/drawings to generate submittal requirements
"""

import json
import os
from pathlib import Path
from typing import Dict, Any, List, Optional
from anthropic import Anthropic

# Initialize Anthropic client
client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))


class SubmittalGenerator:
    """Generates submittal logs from project specifications and drawings"""

    # Standard glazing submittals that apply to most projects
    STANDARD_SUBMITTALS = [
        {"category": "product_data", "description": "Manufacturer's product data for aluminum framing systems", "spec_section": "08 44 00"},
        {"category": "product_data", "description": "Manufacturer's product data for glass types", "spec_section": "08 80 00"},
        {"category": "product_data", "description": "Manufacturer's product data for sealants and glazing compounds", "spec_section": "07 92 00"},
        {"category": "product_data", "description": "Manufacturer's product data for hardware", "spec_section": "08 71 00"},
        {"category": "shop_drawings", "description": "Shop drawings showing elevations, sections, and details", "spec_section": "08 44 00"},
        {"category": "shop_drawings", "description": "Glass setting drawings", "spec_section": "08 80 00"},
        {"category": "samples", "description": "Finish samples for aluminum extrusions", "spec_section": "08 44 00"},
        {"category": "samples", "description": "Glass samples showing color, coating, and thickness", "spec_section": "08 80 00"},
        {"category": "certifications", "description": "Structural calculations sealed by PE", "spec_section": "08 44 00"},
        {"category": "certifications", "description": "Thermal performance calculations (U-factor, SHGC)", "spec_section": "08 44 00"},
        {"category": "test_reports", "description": "Air/water/structural test reports per ASTM standards", "spec_section": "08 44 00"},
        {"category": "warranties", "description": "Manufacturer's warranty for glass coating", "spec_section": "08 80 00"},
        {"category": "warranties", "description": "Installer's warranty for workmanship", "spec_section": "08 44 00"},
        {"category": "maintenance_data", "description": "Maintenance and cleaning instructions", "spec_section": "08 44 00"},
    ]

    def __init__(self):
        self.model = "claude-sonnet-4-20250514"

    def gather_spec_context(self, project_folder: Path) -> Dict[str, Any]:
        """Gather specification and drawing documents for analysis"""
        context = {
            "specs": [],
            "drawings": [],
            "contract": []
        }

        specs_folder = project_folder / "02-Specs"
        drawings_folder = project_folder / "03-Drawings"
        contract_folder = project_folder / "01-Contract"

        # Read spec files (look for text extracts)
        if specs_folder.exists():
            for f in specs_folder.iterdir():
                if f.suffix in ['.txt', '.md', '.json']:
                    try:
                        content = f.read_text(encoding='utf-8', errors='ignore')[:50000]
                        context["specs"].append({
                            "file": f.name,
                            "content": content
                        })
                    except:
                        pass

        # Read drawing notes/extracts
        if drawings_folder.exists():
            for f in drawings_folder.iterdir():
                if f.suffix in ['.txt', '.md', '.json']:
                    try:
                        content = f.read_text(encoding='utf-8', errors='ignore')[:30000]
                        context["drawings"].append({
                            "file": f.name,
                            "content": content
                        })
                    except:
                        pass

        # Read contract documents
        if contract_folder.exists():
            for f in contract_folder.iterdir():
                if f.suffix in ['.txt', '.md', '.json']:
                    try:
                        content = f.read_text(encoding='utf-8', errors='ignore')[:20000]
                        context["contract"].append({
                            "file": f.name,
                            "content": content
                        })
                    except:
                        pass

        return context

    def analyze_for_submittals(self, context: Dict[str, Any], iteration: int = 1) -> List[Dict[str, Any]]:
        """Use AI to analyze documents and extract submittal requirements"""

        # Build prompt with context
        spec_text = "\n\n".join([
            f"=== {s['file']} ===\n{s['content']}"
            for s in context.get("specs", [])
        ]) if context.get("specs") else "No specification documents available"

        drawing_text = "\n\n".join([
            f"=== {d['file']} ===\n{d['content']}"
            for d in context.get("drawings", [])
        ]) if context.get("drawings") else "No drawing extracts available"

        prompt = f"""You are a construction submittal coordinator for a glazing subcontractor.
Analyze the following project documents and identify ALL submittal requirements.

SPECIFICATIONS:
{spec_text}

DRAWING NOTES:
{drawing_text}

For each submittal identified, provide:
1. item_number: A unique identifier (e.g., "GL-001", "GL-002")
2. spec_section: The specification section reference (e.g., "08 44 00")
3. description: Clear description of what needs to be submitted
4. category: One of: product_data, shop_drawings, samples, mock_ups, certifications, warranties, test_reports, maintenance_data, other
5. required: true/false - whether this is explicitly required
6. notes: Any relevant notes about timing, format, or special requirements

This is iteration {iteration} of the analysis. Be thorough and don't miss any submittal requirements.
Focus on:
- Division 08 (Openings) - especially 08 44 00 (Curtainwall), 08 41 00 (Entrances), 08 80 00 (Glazing)
- Division 07 (Thermal/Moisture) - especially 07 92 00 (Sealants)
- Any warranty requirements
- Any test report requirements
- Mock-up requirements
- Sustainable design/LEED requirements

Return ONLY a valid JSON array of submittal objects. No other text."""

        response = client.messages.create(
            model=self.model,
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}]
        )

        # Parse response
        try:
            content = response.content[0].text.strip()
            # Try to extract JSON if wrapped in markdown
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()

            submittals = json.loads(content)
            return submittals if isinstance(submittals, list) else []
        except (json.JSONDecodeError, IndexError):
            return []

    def merge_submittals(self, *submittal_lists: List[Dict]) -> List[Dict[str, Any]]:
        """Merge multiple submittal lists, removing duplicates"""
        seen = set()
        merged = []

        for submittals in submittal_lists:
            for s in submittals:
                # Create a key based on description similarity
                key = (s.get("spec_section", ""), s.get("description", "").lower()[:50])
                if key not in seen:
                    seen.add(key)
                    merged.append(s)

        return merged

    def generate_submittal_log(
        self,
        project_number: str,
        project_folder: Path,
        include_standard: bool = True,
        iterations: int = 2
    ) -> Dict[str, Any]:
        """
        Generate a complete submittal log for a project.

        Args:
            project_number: The project identifier
            project_folder: Path to project folder with documents
            include_standard: Whether to include standard glazing submittals
            iterations: Number of AI analysis passes (for thoroughness)

        Returns:
            Dictionary with submittal log data and metadata
        """
        # Gather document context
        context = self.gather_spec_context(project_folder)

        # Collect all submittals
        all_submittals = []

        # Add standard submittals if requested
        if include_standard:
            for idx, item in enumerate(self.STANDARD_SUBMITTALS, 1):
                all_submittals.append({
                    "item_number": f"STD-{idx:03d}",
                    "spec_section": item["spec_section"],
                    "description": item["description"],
                    "category": item["category"],
                    "required": True,
                    "notes": "Standard glazing submittal",
                    "source": "standard"
                })

        # Run AI analysis iterations
        has_documents = bool(context.get("specs") or context.get("drawings"))

        if has_documents:
            for i in range(iterations):
                ai_submittals = self.analyze_for_submittals(context, iteration=i+1)
                for s in ai_submittals:
                    s["source"] = "ai_analysis"
                all_submittals = self.merge_submittals(all_submittals, ai_submittals)

        # Renumber all submittals
        final_submittals = []
        for idx, s in enumerate(all_submittals, 1):
            final_submittals.append({
                "item_number": f"GL-{idx:03d}",
                "spec_section": s.get("spec_section", ""),
                "description": s.get("description", ""),
                "category": s.get("category", "other"),
                "required": s.get("required", True),
                "status": "not_started",
                "notes": s.get("notes", ""),
                "source": s.get("source", "unknown")
            })

        # Generate summary
        summary = {
            "total_items": len(final_submittals),
            "by_category": {},
            "documents_analyzed": len(context.get("specs", [])) + len(context.get("drawings", [])),
            "iterations": iterations,
            "included_standard": include_standard
        }

        for s in final_submittals:
            cat = s.get("category", "other")
            summary["by_category"][cat] = summary["by_category"].get(cat, 0) + 1

        return {
            "success": True,
            "project_number": project_number,
            "submittals": final_submittals,
            "summary": summary
        }


def generate_submittal_log_excel(submittals: List[Dict], output_path: Path) -> str:
    """Generate an Excel file from submittal data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Submittal Log"

        # Headers
        headers = [
            "Item No.", "Spec Section", "Description", "Category",
            "Required", "Status", "Due Date", "Submitted", "Approved", "Notes"
        ]

        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data rows
        for row_idx, item in enumerate(submittals, 2):
            ws.cell(row=row_idx, column=1, value=item.get("item_number", ""))
            ws.cell(row=row_idx, column=2, value=item.get("spec_section", ""))
            ws.cell(row=row_idx, column=3, value=item.get("description", ""))
            ws.cell(row=row_idx, column=4, value=item.get("category", "").replace("_", " ").title())
            ws.cell(row=row_idx, column=5, value="Yes" if item.get("required") else "No")
            ws.cell(row=row_idx, column=6, value=item.get("status", "Not Started").replace("_", " ").title())
            ws.cell(row=row_idx, column=7, value="")  # Due date
            ws.cell(row=row_idx, column=8, value="")  # Submitted date
            ws.cell(row=row_idx, column=9, value="")  # Approved date
            ws.cell(row=row_idx, column=10, value=item.get("notes", ""))

            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 30

        # Freeze header row
        ws.freeze_panes = 'A2'

        wb.save(output_path)
        return str(output_path)

    except Exception as e:
        raise Exception(f"Failed to generate Excel: {str(e)}")


# Standalone test
if __name__ == "__main__":
    generator = SubmittalGenerator()
    print("Submittal Generator initialized")
    print(f"Standard submittals count: {len(generator.STANDARD_SUBMITTALS)}")
