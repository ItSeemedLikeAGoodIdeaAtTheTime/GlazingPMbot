"""
Template Processor - Excel Template Filler

Reads user-provided Excel templates (SOV, Budget) and fills them with project data.
Uses Claude to intelligently map project data to template cells.

Flow:
1. Read the template Excel file and extract its structure
2. Get project data (contract analysis, scope, etc.)
3. Use Claude to determine which cells should contain what data
4. Fill in the template and save as a new file
"""

import os
import sys
import json
import copy
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple
from anthropic import Anthropic

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class TemplateProcessor:
    """Processes Excel templates and fills them with project data"""

    def __init__(self, api_key: Optional[str] = None):
        """Initialize with Anthropic API key"""
        self.api_key = api_key or os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("anthropicAPIkey")
        if not self.api_key:
            raise ValueError("Anthropic API key not found")

        self.client = Anthropic(api_key=self.api_key)

    def read_template_structure(self, template_path: Path) -> Dict[str, Any]:
        """
        Read an Excel template and extract its structure.
        Returns info about sheets, headers, and cell layout.
        """
        wb = load_workbook(template_path, data_only=False)

        structure = {
            "file_name": template_path.name,
            "sheets": []
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            sheet_info = {
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "cells": []
            }

            # Extract cell contents (first 50 rows, 20 columns for analysis)
            for row in range(1, min(ws.max_row + 1, 51)):
                for col in range(1, min(ws.max_column + 1, 21)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        sheet_info["cells"].append({
                            "cell": f"{get_column_letter(col)}{row}",
                            "row": row,
                            "col": col,
                            "value": str(cell.value)[:100],  # Truncate long values
                            "is_formula": str(cell.value).startswith("=") if cell.value else False
                        })

            structure["sheets"].append(sheet_info)

        wb.close()
        return structure

    def analyze_template_for_mapping(
        self,
        template_structure: Dict[str, Any],
        project_data: Dict[str, Any],
        template_type: str = "sov"
    ) -> Dict[str, Any]:
        """
        Use Claude to analyze the template and determine cell mappings.
        Returns a mapping of cell addresses to values.
        """

        prompt = f"""You are analyzing an Excel template for a commercial glazing project.

TEMPLATE TYPE: {template_type.upper()} ({"Schedule of Values" if template_type == "sov" else "Internal Budget"})

TEMPLATE STRUCTURE:
{json.dumps(template_structure, indent=2)}

PROJECT DATA:
{json.dumps(project_data, indent=2)}

Your task:
1. Analyze the template structure to understand what data goes where
2. Map the project data to the appropriate cells
3. Return a JSON object with cell mappings

RULES:
- Do NOT overwrite cells that contain formulas (is_formula: true)
- Match data intelligently (e.g., "Project Name" header -> put project name in adjacent cell)
- For SOV templates: fill in line items with descriptions, spec sections, and values
- For Budget templates: fill in cost codes, quantities, unit costs, totals
- Leave cells blank if no appropriate data exists
- Be conservative - only fill cells where you're confident about the mapping

Return a JSON object with this structure:
{{
    "sheet_name": "Sheet1",
    "mappings": [
        {{
            "cell": "B2",
            "value": "Example Project Name",
            "reason": "Cell next to 'Project Name' label"
        }},
        {{
            "cell": "C5",
            "value": 125000.00,
            "reason": "Contract value field"
        }}
    ],
    "line_items": {{
        "start_row": 10,
        "columns": {{
            "item_number": "A",
            "description": "B",
            "spec_section": "C",
            "value": "D"
        }},
        "items": [
            {{
                "item_number": 1,
                "description": "General Conditions",
                "spec_section": "088000",
                "value": 15000.00
            }}
        ]
    }},
    "notes": "Any observations about the template or data mapping"
}}

If no line items section exists, omit the "line_items" field.
Return ONLY valid JSON, no markdown formatting."""

        response = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8000,
            messages=[{
                "role": "user",
                "content": prompt
            }]
        )

        response_text = response.content[0].text

        # Extract JSON from response
        if '```json' in response_text:
            start = response_text.find('```json') + 7
            end = response_text.find('```', start)
            response_text = response_text[start:end].strip()
        elif '```' in response_text:
            start = response_text.find('```') + 3
            end = response_text.find('```', start)
            response_text = response_text[start:end].strip()

        try:
            mapping = json.loads(response_text)
            return mapping
        except json.JSONDecodeError as e:
            print(f"WARNING: Could not parse mapping JSON: {e}")
            return {"mappings": [], "notes": f"Parse error: {e}"}

    def fill_template(
        self,
        template_path: Path,
        output_path: Path,
        mapping: Dict[str, Any]
    ) -> bool:
        """
        Fill an Excel template with the provided mapping.
        Returns True on success.
        """
        wb = load_workbook(template_path)

        sheet_name = mapping.get("sheet_name", wb.sheetnames[0])
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # Apply individual cell mappings
        for cell_mapping in mapping.get("mappings", []):
            cell_addr = cell_mapping.get("cell")
            value = cell_mapping.get("value")

            if cell_addr and value is not None:
                try:
                    ws[cell_addr] = value
                except Exception as e:
                    print(f"WARNING: Could not set {cell_addr}: {e}")

        # Apply line items if present
        line_items = mapping.get("line_items")
        if line_items:
            start_row = line_items.get("start_row", 10)
            columns = line_items.get("columns", {})
            items = line_items.get("items", [])

            for i, item in enumerate(items):
                row = start_row + i

                for field, col_letter in columns.items():
                    if field in item and item[field] is not None:
                        try:
                            ws[f"{col_letter}{row}"] = item[field]
                        except Exception as e:
                            print(f"WARNING: Could not set {col_letter}{row}: {e}")

        # Save the filled template
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        wb.close()

        return True

    def process_template(
        self,
        template_path: Path,
        project_number: str,
        project_data: Dict[str, Any],
        template_type: str = "sov",
        output_dir: Optional[Path] = None
    ) -> Dict[str, Any]:
        """
        Main entry point: process a template and fill it with project data.

        Args:
            template_path: Path to the Excel template
            project_number: Project identifier (e.g., "P001")
            project_data: Dict containing contract_analysis, scope_analysis, etc.
            template_type: "sov" or "budget"
            output_dir: Where to save the filled template (default: Output/Filled_Templates)

        Returns:
            Dict with success status, output path, and any notes
        """
        print(f"\n{'='*60}")
        print(f"  TEMPLATE PROCESSOR: {template_type.upper()}")
        print(f"  Project: {project_number}")
        print(f"{'='*60}\n")

        if not template_path.exists():
            return {
                "success": False,
                "error": f"Template not found: {template_path}"
            }

        # Step 1: Read template structure
        print("[1/4] Reading template structure...")
        try:
            structure = self.read_template_structure(template_path)
            print(f"  Found {len(structure['sheets'])} sheet(s)")
            for sheet in structure['sheets']:
                print(f"    - {sheet['name']}: {len(sheet['cells'])} cells with data")
        except Exception as e:
            return {"success": False, "error": f"Failed to read template: {e}"}

        # Step 2: Analyze and create mapping
        print("[2/4] Analyzing template and mapping data...")
        try:
            mapping = self.analyze_template_for_mapping(structure, project_data, template_type)
            mapping_count = len(mapping.get("mappings", []))
            line_item_count = len(mapping.get("line_items", {}).get("items", []))
            print(f"  Created {mapping_count} cell mappings")
            if line_item_count:
                print(f"  Created {line_item_count} line items")
        except Exception as e:
            return {"success": False, "error": f"Failed to analyze template: {e}"}

        # Step 3: Fill template
        print("[3/4] Filling template...")
        if output_dir is None:
            output_dir = Path("Output/Filled_Templates")

        output_filename = f"{project_number}_{template_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = output_dir / output_filename

        try:
            self.fill_template(template_path, output_path, mapping)
            print(f"  Saved to: {output_path}")
        except Exception as e:
            return {"success": False, "error": f"Failed to fill template: {e}"}

        # Step 4: Save mapping for reference
        print("[4/4] Saving mapping reference...")
        mapping_path = output_dir / f"{project_number}_{template_type}_mapping.json"
        try:
            with open(mapping_path, 'w', encoding='utf-8') as f:
                json.dump(mapping, f, indent=2)
        except Exception as e:
            print(f"  WARNING: Could not save mapping: {e}")

        print(f"\n{'='*60}")
        print(f"  TEMPLATE PROCESSING COMPLETE")
        print(f"{'='*60}")
        print(f"\nOutput: {output_path}")
        if mapping.get("notes"):
            print(f"Notes: {mapping['notes']}")

        return {
            "success": True,
            "output_path": str(output_path),
            "mapping_path": str(mapping_path),
            "mappings_applied": mapping_count,
            "line_items_added": line_item_count,
            "notes": mapping.get("notes")
        }


def main():
    """CLI interface for template processor"""
    if len(sys.argv) < 3:
        print("\nUsage: python template_processor.py <project_number> <template_type>")
        print("\nArguments:")
        print("  project_number  - Project ID (e.g., P001)")
        print("  template_type   - 'sov' or 'budget'")
        print("\nExample:")
        print("  python template_processor.py P001 sov")
        print("  python template_processor.py P001 budget")
        print("\nThe script will look for templates in:")
        print("  Projects/{project_number}-*/06-Templates/")
        sys.exit(1)

    project_number = sys.argv[1]
    template_type = sys.argv[2].lower()

    if template_type not in ["sov", "budget"]:
        print(f"ERROR: template_type must be 'sov' or 'budget', got '{template_type}'")
        sys.exit(1)

    # Find project folder
    project_folders = list(Path("Projects").glob(f"{project_number}-*"))
    if not project_folders:
        print(f"ERROR: No project found with number: {project_number}")
        sys.exit(1)

    project_folder = project_folders[0]
    templates_folder = project_folder / "06-Templates"

    # Find template file
    template_pattern = f"*{template_type}*" if template_type else "*.xlsx"
    template_files = list(templates_folder.glob("*.xlsx")) + list(templates_folder.glob("*.xls"))

    # Filter to matching template type
    matching_templates = [
        f for f in template_files
        if template_type in f.name.lower()
    ]

    if not matching_templates:
        # Fall back to any Excel file
        matching_templates = template_files

    if not matching_templates:
        print(f"ERROR: No {template_type} template found in {templates_folder}")
        print(f"Upload a template file to: {templates_folder}")
        sys.exit(1)

    template_path = matching_templates[0]
    print(f"Using template: {template_path.name}")

    # Load project data
    project_data = {}

    # Load contract analysis
    analysis_file = Path(f"Output/Reports/{project_number}_contract_analysis.json")
    if analysis_file.exists():
        with open(analysis_file, 'r', encoding='utf-8') as f:
            project_data["contract_analysis"] = json.load(f)
    else:
        print(f"WARNING: Contract analysis not found: {analysis_file}")

    # Load scope analysis
    scope_file = Path(f"Output/Scope_Analysis/{project_number}_scope_analysis.json")
    if scope_file.exists():
        with open(scope_file, 'r', encoding='utf-8') as f:
            project_data["scope_analysis"] = json.load(f)

    # Load budget if exists
    budget_file = Path(f"Output/Budgets/{project_number}_internal_budget.json")
    if budget_file.exists():
        with open(budget_file, 'r', encoding='utf-8') as f:
            project_data["budget"] = json.load(f)

    # Load SOV if exists
    sov_file = Path(f"Output/Draft_SOV/{project_number}_SOV.json")
    if sov_file.exists():
        with open(sov_file, 'r', encoding='utf-8') as f:
            project_data["sov"] = json.load(f)

    if not project_data:
        print("ERROR: No project data found. Run contract analysis first:")
        print(f"  python scripts/contract_processor.py {project_number}")
        sys.exit(1)

    # Process template
    try:
        processor = TemplateProcessor()
        result = processor.process_template(
            template_path=template_path,
            project_number=project_number,
            project_data=project_data,
            template_type=template_type
        )

        if result["success"]:
            print("\n[OK] Template processing complete!")
            sys.exit(0)
        else:
            print(f"\n[ERROR] {result.get('error')}")
            sys.exit(1)

    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
