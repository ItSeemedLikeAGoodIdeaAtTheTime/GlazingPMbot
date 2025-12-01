"""
AI Estimator - Intelligent Value Estimation Engine

Uses all available project context to estimate realistic values:
- Contract documents and analysis
- Vendor quotes
- Invoices and actuals
- Previous billings (for SOV chain)
- Historical project data
- Industry knowledge

Two modes:
1. Budget Mode: One-time cost estimate with cost codes
2. SOV Mode: Monthly billing chain with previous billing constraints
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List
from anthropic import Anthropic

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("WARNING: openpyxl not installed. Excel features disabled.")


class AIEstimator:
    """AI-powered estimation engine for glazing projects"""

    def __init__(self, api_key: Optional[str] = None):
        self.api_key = api_key or os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("anthropicAPIkey")
        if not self.api_key:
            raise ValueError("Anthropic API key not found")

        self.client = Anthropic(api_key=self.api_key)

    def gather_project_context(self, project_number: str, project_folder: Path) -> Dict[str, Any]:
        """
        Gather all available context for a project.
        Returns everything we know that could inform estimates.
        """
        context = {
            "project_number": project_number,
            "gathered_at": datetime.now().isoformat(),
            "contract_analysis": None,
            "scope_analysis": None,
            "vendor_quotes": [],
            "invoices": [],
            "previous_billings": [],
            "existing_budget": None,
            "existing_sov": None,
            "project_info": None
        }

        # Load contract analysis
        analysis_file = Path(f"Output/Reports/{project_number}_contract_analysis.json")
        if analysis_file.exists():
            with open(analysis_file, 'r', encoding='utf-8') as f:
                context["contract_analysis"] = json.load(f)

        # Load scope analysis
        scope_file = Path(f"Output/Scope_Analysis/{project_number}_scope_analysis.json")
        if scope_file.exists():
            with open(scope_file, 'r', encoding='utf-8') as f:
                context["scope_analysis"] = json.load(f)

        # Load existing budget if any
        budget_file = Path(f"Output/Budgets/{project_number}_internal_budget.json")
        if budget_file.exists():
            with open(budget_file, 'r', encoding='utf-8') as f:
                context["existing_budget"] = json.load(f)

        # Load existing SOV if any
        sov_file = Path(f"Output/Draft_SOV/{project_number}_SOV.json")
        if sov_file.exists():
            with open(sov_file, 'r', encoding='utf-8') as f:
                context["existing_sov"] = json.load(f)

        # Load project info
        project_info_file = project_folder / "project_info.json"
        if project_info_file.exists():
            with open(project_info_file, 'r', encoding='utf-8') as f:
                context["project_info"] = json.load(f)

        # Load vendor quotes from documents table or folder
        quotes_folder = project_folder / "vendor_quotes"
        if quotes_folder.exists():
            for quote_file in quotes_folder.glob("*.json"):
                with open(quote_file, 'r', encoding='utf-8') as f:
                    context["vendor_quotes"].append(json.load(f))

        # Load invoices
        invoices_folder = project_folder / "invoices"
        if invoices_folder.exists():
            for invoice_file in invoices_folder.glob("*.json"):
                with open(invoice_file, 'r', encoding='utf-8') as f:
                    context["invoices"].append(json.load(f))

        # Load previous billings (SOV history)
        billings_folder = project_folder / "billings"
        if billings_folder.exists():
            for billing_file in sorted(billings_folder.glob("*.json")):
                with open(billing_file, 'r', encoding='utf-8') as f:
                    context["previous_billings"].append(json.load(f))

        return context

    def read_template_structure(self, template_path: Path) -> Dict[str, Any]:
        """Extract structure from Excel template"""
        wb = load_workbook(template_path, data_only=False)

        structure = {
            "file_name": template_path.name,
            "sheets": []
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            sheet_info = {
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "cells": []
            }

            # Extract cell contents (first 100 rows, 30 columns)
            for row in range(1, min(ws.max_row + 1, 101)):
                for col in range(1, min(ws.max_column + 1, 31)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        sheet_info["cells"].append({
                            "cell": f"{get_column_letter(col)}{row}",
                            "row": row,
                            "col": col,
                            "value": cell_value[:200],
                            "is_formula": cell_value.startswith("=") if cell_value else False
                        })

            structure["sheets"].append(sheet_info)

        wb.close()
        return structure

    def generate_budget(
        self,
        project_number: str,
        project_folder: Path,
        template_path: Optional[Path] = None,
        revision: int = 1
    ) -> Dict[str, Any]:
        """
        Generate internal budget using AI estimation.

        Args:
            project_number: Project ID
            project_folder: Path to project folder
            template_path: Optional Excel template for cost code structure
            revision: Budget revision number

        Returns:
            Dict with budget data and output paths
        """
        print(f"\n{'='*60}")
        print(f"  AI BUDGET ESTIMATION: {project_number} (Rev {revision})")
        print(f"{'='*60}\n")

        # Gather all context
        print("[1/4] Gathering project context...")
        context = self.gather_project_context(project_number, project_folder)

        context_summary = []
        if context["contract_analysis"]:
            context_summary.append("contract analysis")
        if context["scope_analysis"]:
            context_summary.append("scope analysis")
        if context["vendor_quotes"]:
            context_summary.append(f"{len(context['vendor_quotes'])} vendor quotes")
        if context["invoices"]:
            context_summary.append(f"{len(context['invoices'])} invoices")

        print(f"  Available: {', '.join(context_summary) if context_summary else 'minimal data'}")

        # Read template if provided
        template_structure = None
        if template_path and template_path.exists():
            print(f"[2/4] Reading budget template: {template_path.name}")
            template_structure = self.read_template_structure(template_path)
        else:
            print("[2/4] No template provided - using standard cost codes")

        # Build estimation prompt
        print("[3/4] AI estimating budget values...")

        prompt = self._build_budget_prompt(context, template_structure, revision)

        response = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}]
        )

        response_text = response.content[0].text

        # Parse JSON response
        budget_data = self._parse_json_response(response_text)

        if not budget_data:
            return {"success": False, "error": "Failed to parse budget response"}

        # Add metadata
        budget_data["metadata"] = {
            "project_number": project_number,
            "revision": revision,
            "generated_at": datetime.now().isoformat(),
            "context_used": context_summary
        }

        # Save outputs
        print("[4/4] Saving budget...")
        output_dir = Path("Output/Budgets")
        output_dir.mkdir(parents=True, exist_ok=True)

        # JSON output
        json_file = output_dir / f"{project_number}_budget_rev{revision}.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(budget_data, f, indent=2)

        # Fill Excel template if provided
        excel_file = None
        if template_path and template_path.exists():
            excel_file = self._fill_budget_template(
                template_path,
                budget_data,
                project_number,
                revision
            )

        print(f"\n{'='*60}")
        print(f"  BUDGET COMPLETE: Rev {revision}")
        print(f"{'='*60}")
        print(f"  JSON: {json_file}")
        if excel_file:
            print(f"  Excel: {excel_file}")

        return {
            "success": True,
            "project_number": project_number,
            "revision": revision,
            "budget_data": budget_data,
            "json_file": str(json_file),
            "excel_file": str(excel_file) if excel_file else None
        }

    def generate_sov(
        self,
        project_number: str,
        project_folder: Path,
        billing_month: str,
        billing_year: int,
        template_path: Optional[Path] = None
    ) -> Dict[str, Any]:
        """
        Generate Schedule of Values for a specific month.

        Args:
            project_number: Project ID
            project_folder: Path to project folder
            billing_month: Month name (e.g., "September")
            billing_year: Year (e.g., 2024)
            template_path: Optional Excel template

        Returns:
            Dict with SOV data and output paths
        """
        print(f"\n{'='*60}")
        print(f"  AI SOV ESTIMATION: {project_number}")
        print(f"  Billing Period: {billing_month} {billing_year}")
        print(f"{'='*60}\n")

        # Gather all context
        print("[1/5] Gathering project context...")
        context = self.gather_project_context(project_number, project_folder)

        context_summary = []
        if context["contract_analysis"]:
            context_summary.append("contract analysis")
        if context["vendor_quotes"]:
            context_summary.append(f"{len(context['vendor_quotes'])} quotes")
        if context["invoices"]:
            context_summary.append(f"{len(context['invoices'])} invoices")
        if context["previous_billings"]:
            context_summary.append(f"{len(context['previous_billings'])} previous billings")

        print(f"  Available: {', '.join(context_summary) if context_summary else 'minimal data'}")

        # Find the most recent previous billing
        previous_billing = None
        if context["previous_billings"]:
            # Sort by date and get most recent
            sorted_billings = sorted(
                context["previous_billings"],
                key=lambda x: (x.get("year", 0), self._month_to_num(x.get("month", "")))
            )
            previous_billing = sorted_billings[-1] if sorted_billings else None
            if previous_billing:
                print(f"  Previous billing: {previous_billing.get('month')} {previous_billing.get('year')}")

        # Read template if provided
        template_structure = None
        if template_path and template_path.exists():
            print(f"[2/5] Reading SOV template: {template_path.name}")
            template_structure = self.read_template_structure(template_path)
        else:
            print("[2/5] No template provided - using standard SOV format")

        # Build estimation prompt
        print("[3/5] AI estimating SOV values...")

        prompt = self._build_sov_prompt(
            context,
            template_structure,
            billing_month,
            billing_year,
            previous_billing
        )

        response = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}]
        )

        response_text = response.content[0].text
        sov_data = self._parse_json_response(response_text)

        if not sov_data:
            return {"success": False, "error": "Failed to parse SOV response"}

        # Add metadata
        sov_data["metadata"] = {
            "project_number": project_number,
            "billing_month": billing_month,
            "billing_year": billing_year,
            "generated_at": datetime.now().isoformat(),
            "context_used": context_summary,
            "previous_billing": f"{previous_billing.get('month')} {previous_billing.get('year')}" if previous_billing else None
        }

        # Save outputs
        print("[4/5] Saving SOV...")
        output_dir = Path("Output/SOV")
        output_dir.mkdir(parents=True, exist_ok=True)

        # JSON output
        json_file = output_dir / f"{project_number}_SOV_{billing_year}_{billing_month}.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(sov_data, f, indent=2)

        # Also save to project billings folder for chain tracking
        print("[5/5] Recording in billing history...")
        billings_folder = project_folder / "billings"
        billings_folder.mkdir(parents=True, exist_ok=True)

        billing_record = {
            "month": billing_month,
            "year": billing_year,
            "sov_data": sov_data,
            "generated_at": datetime.now().isoformat()
        }

        billing_file = billings_folder / f"{billing_year}_{self._month_to_num(billing_month):02d}_{billing_month}.json"
        with open(billing_file, 'w', encoding='utf-8') as f:
            json.dump(billing_record, f, indent=2)

        # Fill Excel template if provided
        excel_file = None
        if template_path and template_path.exists():
            excel_file = self._fill_sov_template(
                template_path,
                sov_data,
                project_number,
                billing_month,
                billing_year
            )

        print(f"\n{'='*60}")
        print(f"  SOV COMPLETE: {billing_month} {billing_year}")
        print(f"{'='*60}")
        print(f"  JSON: {json_file}")
        if excel_file:
            print(f"  Excel: {excel_file}")

        return {
            "success": True,
            "project_number": project_number,
            "billing_month": billing_month,
            "billing_year": billing_year,
            "sov_data": sov_data,
            "json_file": str(json_file),
            "excel_file": str(excel_file) if excel_file else None
        }

    def _build_budget_prompt(
        self,
        context: Dict[str, Any],
        template_structure: Optional[Dict[str, Any]],
        revision: int
    ) -> str:
        """Build the prompt for budget estimation"""

        prompt = f"""You are an expert commercial glazing estimator creating an INTERNAL BUDGET.

This is Revision {revision} of the budget for this project.

PROJECT CONTEXT:
{json.dumps(context, indent=2, default=str)}

"""
        if template_structure:
            prompt += f"""
BUDGET TEMPLATE STRUCTURE:
The client has provided this Excel template. You must fill values that match this structure.
{json.dumps(template_structure, indent=2)}

Analyze the template to understand:
- What cost codes/categories are expected
- What columns need values (quantity, unit cost, total, etc.)
- Any formulas that will auto-calculate (don't fill those)
"""
        else:
            prompt += """
No template provided. Use standard glazing cost code structure:
- Labor (L-): Glaziers, Caulkers, Shop labor, Supervision
- Materials (M-): Glass, Framing, Hardware, Sealants
- Indirect: Drive time, Shipping, Equipment, Misc
"""

        prompt += """

YOUR TASK:
1. Analyze ALL available context (contract, quotes, invoices, scope)
2. Estimate realistic costs for each line item
3. If vendor quotes exist, use those actual numbers
4. If no quotes, estimate based on industry knowledge and project scope
5. Allocate contract value intelligently across categories
6. Be conservative on labor estimates (glazing is labor-intensive)

ESTIMATION GUIDELINES:
- Labor: $55-85/hr depending on skill level
- Glass: $25-60/sqft depending on type (monolithic vs IGU vs specialty)
- Aluminum framing: $15-35/lnft depending on system
- Hardware: Count doors, price per door type
- Sealants: $8-15/lnft for perimeter, $3-8/lnft for interior
- Add 10-15% contingency for unknowns

Return JSON:
{
    "summary": {
        "total_labor": 0.00,
        "total_materials": 0.00,
        "total_indirect": 0.00,
        "contingency": 0.00,
        "grand_total": 0.00
    },
    "line_items": [
        {
            "cost_code": "L-GLZ-001",
            "category": "Labor",
            "description": "Glaziers - Field Installation",
            "quantity": 0,
            "unit": "hours",
            "unit_cost": 0.00,
            "total": 0.00,
            "basis": "How you estimated this"
        }
    ],
    "assumptions": ["List key assumptions"],
    "risks": ["List cost risks identified"],
    "template_mappings": [
        {"cell": "B5", "value": "Project Name Here", "field": "project_name"}
    ]
}

If a template was provided, include template_mappings showing which cells get which values.
Return ONLY valid JSON."""

        return prompt

    def _build_sov_prompt(
        self,
        context: Dict[str, Any],
        template_structure: Optional[Dict[str, Any]],
        billing_month: str,
        billing_year: int,
        previous_billing: Optional[Dict[str, Any]]
    ) -> str:
        """Build the prompt for SOV estimation"""

        prompt = f"""You are an expert commercial glazing project manager creating a Schedule of Values (SOV).

BILLING PERIOD: {billing_month} {billing_year}

PROJECT CONTEXT:
{json.dumps(context, indent=2, default=str)}

"""
        if previous_billing:
            prompt += f"""
CRITICAL - PREVIOUS BILLING EXISTS:
{json.dumps(previous_billing, indent=2, default=str)}

You MUST:
1. Use the EXACT SAME line items and structure as previous billing
2. Keep "Scheduled Value" column identical (these are locked)
3. Only update: "Work Completed This Period", "Materials Stored", "Total Completed"
4. Calculate new "Balance to Finish" and "% Complete"
5. Never exceed the Scheduled Value for any line item
"""
        else:
            prompt += """
This is the FIRST billing for this project.
You are establishing the SOV structure that all future billings will follow.
Choose line items carefully - they become locked for the project duration.
"""

        if template_structure:
            prompt += f"""
SOV TEMPLATE STRUCTURE:
{json.dumps(template_structure, indent=2)}

Fill values according to this template's column structure.
"""

        prompt += """

SOV STRUCTURE (AIA G702/G703 style):
- Item Number
- Description of Work
- Scheduled Value (total contract value for this item)
- Work Completed - Previous (from prior billings)
- Work Completed - This Period (what we're billing now)
- Materials Stored (materials on site not yet installed)
- Total Completed & Stored
- % Complete
- Balance to Finish
- Retainage (typically 5-10%)

BILLING STRATEGY:
- Front-load general conditions and submittals (bill early)
- Bill for materials when purchased/delivered
- Bill labor progressively as work completes
- Be aggressive but defensible
- Match billings to actual progress + slight optimism

Return JSON:
{
    "application_number": 1,
    "billing_period": {"month": "September", "year": 2024},
    "contract_value": 0.00,
    "summary": {
        "original_contract": 0.00,
        "change_orders": 0.00,
        "revised_contract": 0.00,
        "total_completed": 0.00,
        "retainage": 0.00,
        "total_earned_less_retainage": 0.00,
        "previous_payments": 0.00,
        "current_payment_due": 0.00
    },
    "line_items": [
        {
            "item_number": 1,
            "description": "General Conditions / Submittals",
            "spec_section": "088000",
            "scheduled_value": 0.00,
            "previous_completed": 0.00,
            "this_period": 0.00,
            "materials_stored": 0.00,
            "total_completed": 0.00,
            "percent_complete": 0.0,
            "balance_to_finish": 0.00,
            "retainage": 0.00
        }
    ],
    "template_mappings": [
        {"cell": "B5", "value": "...", "field": "project_name"}
    ],
    "notes": "Any notes about this billing"
}

Return ONLY valid JSON."""

        return prompt

    def _parse_json_response(self, response_text: str) -> Optional[Dict[str, Any]]:
        """Parse JSON from Claude response"""
        # Try to extract JSON from markdown blocks
        if '```json' in response_text:
            start = response_text.find('```json') + 7
            end = response_text.find('```', start)
            response_text = response_text[start:end].strip()
        elif '```' in response_text:
            start = response_text.find('```') + 3
            end = response_text.find('```', start)
            response_text = response_text[start:end].strip()

        try:
            return json.loads(response_text)
        except json.JSONDecodeError as e:
            print(f"WARNING: JSON parse error: {e}")
            print(f"Response preview: {response_text[:500]}...")
            return None

    def _month_to_num(self, month: str) -> int:
        """Convert month name to number"""
        months = {
            "january": 1, "february": 2, "march": 3, "april": 4,
            "may": 5, "june": 6, "july": 7, "august": 8,
            "september": 9, "october": 10, "november": 11, "december": 12
        }
        return months.get(month.lower(), 0)

    def _fill_budget_template(
        self,
        template_path: Path,
        budget_data: Dict[str, Any],
        project_number: str,
        revision: int
    ) -> Optional[Path]:
        """Fill Excel budget template with estimated values"""
        try:
            wb = load_workbook(template_path)
            ws = wb.active

            # Apply template mappings from AI response
            for mapping in budget_data.get("template_mappings", []):
                cell = mapping.get("cell")
                value = mapping.get("value")
                if cell and value is not None:
                    try:
                        ws[cell] = value
                    except Exception as e:
                        print(f"  Warning: Could not set {cell}: {e}")

            # Save filled template
            output_dir = Path("Output/Filled_Templates")
            output_dir.mkdir(parents=True, exist_ok=True)

            output_path = output_dir / f"{project_number}_budget_rev{revision}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(output_path)
            wb.close()

            return output_path
        except Exception as e:
            print(f"  Warning: Could not fill budget template: {e}")
            return None

    def _fill_sov_template(
        self,
        template_path: Path,
        sov_data: Dict[str, Any],
        project_number: str,
        billing_month: str,
        billing_year: int
    ) -> Optional[Path]:
        """Fill Excel SOV template with estimated values"""
        try:
            wb = load_workbook(template_path)
            ws = wb.active

            # Apply template mappings from AI response
            for mapping in sov_data.get("template_mappings", []):
                cell = mapping.get("cell")
                value = mapping.get("value")
                if cell and value is not None:
                    try:
                        ws[cell] = value
                    except Exception as e:
                        print(f"  Warning: Could not set {cell}: {e}")

            # Save filled template
            output_dir = Path("Output/Filled_Templates")
            output_dir.mkdir(parents=True, exist_ok=True)

            output_path = output_dir / f"{project_number}_SOV_{billing_year}_{billing_month}_{datetime.now().strftime('%H%M%S')}.xlsx"
            wb.save(output_path)
            wb.close()

            return output_path
        except Exception as e:
            print(f"  Warning: Could not fill SOV template: {e}")
            return None


def main():
    """CLI interface"""
    print("\nAI Estimator - Use via API or import as module")
    print("\nExample usage:")
    print("  from ai_estimator import AIEstimator")
    print("  estimator = AIEstimator()")
    print("  result = estimator.generate_budget('P001', Path('Projects/P001-Test'))")
    print("  result = estimator.generate_sov('P001', Path('Projects/P001-Test'), 'September', 2024)")


if __name__ == "__main__":
    main()
